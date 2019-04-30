def main():
    pass

if __name__ == '__main__':
    main()


import arcpy, openpyxl, sys, os
from openpyxl.reader.excel import load_workbook

arcpy.AddMessage("Locating Relevant File Paths...:")
scriptPath = sys.path[0]
arcpy.AddMessage("Script Folder: " + scriptPath)

scratchData = arcpy.GetParameterAsText(4)
if scratchData == "#" or not scratchData:
    scratchData = r"E:\gis\projects\IL\Chicago\CDOT_Menu_Pro\Spatial\gdb\working\scratch\CIC_Excel_Scratch"
    arcpy.AddMessage("Data Folder: " + scratchData)

templatePath = arcpy.GetParameterAsText(5)
if templatePath == "#" or not templatePath:
    templatePath = r"E:\gis\projects\IL\Chicago\CDOT_Menu_Pro\Scripts\ArcToolbox\Toolset\Templates"
    arcpy.AddMessage("Template Folder: " + templatePath)

cce = arcpy.GetParameterAsText(0)
if cce == "#" or not cce:
    cce = "05-ASlRS-703"
blockNo = arcpy.GetParameterAsText(1)
if blockNo == "#" or not blockNo:
    blockNo = "1"

outputWB = arcpy.GetParameterAsText(2)
if outputWB == "#" or not outputWB:
    outputWB = r"E:\gis\projects\IL\Chicago\CDOT_Menu_Pro\Map_Docs\Final\Summary Sheets"
    arcpy.AddMessage("Aquiring User Parameters..." + "CCE " + cce + " and block Number " + blockNo)

whereClause = '"strCCEProjectNumber"' + " = '" + str(cce) + "' AND " + '"BlockNo"' + " = " + blockNo
arcpy.AddMessage(whereClause)

#Code to iterate throught the calc table and extract the values
dataset_name = arcpy.GetParameterAsText(3)
if dataset_name == '#' or not dataset_name:
    dataset_name = r"E:\gis\projects\IL\Chicago\CDOT_Menu_Pro\Spatial\gdb\Lex_CDOT_MapOutput.sde\Lex_CDOT_MapOutput.DBO.Calcs\Lex_CDOT_MapOutput.DBO.Blocks_FinalCalcs"

def header_and_iterator(dataset_name):
    data_description = arcpy.Describe(dataset_name)
    # Populates a list of field names if they are not of Geometry, Raster, or Blob type
    fieldnames = [f.name for f in data_description.fields if f.name not in ["Geometry", "Raster", "Blob"]]
    print fieldnames

    fieldnamesred = []
    for n in fieldnames:
        if n in ["PageNo", "BlockNo", 'lngProjectID',
'strCCEProjectNumber',
'SurveyBy',
'SurveyByDate',
'CalcBy',
'CalcByDate',
'strBlockLocation',
'strBlockTo',
'strBlockFrom',
'CICEarthExec_CoreCY',
'CICPaveRemove_CoreSY',
'CICDrvAlleyPaveRemove_CoreSY',
'CICT4CurbRemove_CoreFT',
'CICCurbGutRemove_CoreFT',
'CICSidewalkRemove_CoreSF',
'CICPCCSW5inGreater1800_CoreSF',
'CICSubGranMatTypeB_CoreCY',
'CICPCCBaseCourse10in_CoreSY',
'CICPCCBaseCourse12in_CoreSY',
'CICAlleyHESPCC8in_CoreSY',
'CICDriveHESPCC8in_CoreSY',
'CICSidewalkHESPCC8in_CoreSY',
'CICPCCSW5inLess1800_CoreSF',
'CICBitumMatPrimeCoat_CoreGAL',
'CICCurbGutLess300_CoreLFT',
'CICCurbGutGreater300_CoreLFT',
'CICDrillGroutTieBars_CoreEA',
'CICVertCurbT4_CoreLFT',
'CICCurbGutT3_CoreLFT',
'CICDrainUtilityAdj_CoreEA',
'CICWaterShutRem_CoreEA',
'CICCatchBasinMHEtc_CoreEA',
'CICTreeProtection_CoreEA',
'CICPulvTopsoilMix_CoreCY',
'CICRootPruning_CoreEA',
'CICShredHardwoodBark_CoreCY',
'CICHydraulicSeed_CoreSY',
'CICCrushedStone_CoreTON',
'CICThermoPave6in_CoreLFT',
'CICThermoPave24in_CoreLFT',
'CICPaveMarkRemove_CoreSF',
'CICThermoPave4in_CoreLFT',
'CICHMAN30Type1_CoreSY',
'CICHMAN30Type2_CoreSY',
'CICHMAN70Type1_CoreSY',
'CICHMAN70Type2_CoreSY',
'CICEarthExec_ADA_CY',
'CICT4CurbRemove_ADA_FT',
'CICCurbGutRemove_ADA_FT',
'CICSidewalkRemove_ADA_SF',
'CICSubGranMatTypeB_ADA_CY',
'CICPCCBaseCourse7in_ADA_SY',
'CICPCCBaseCourse9in_ADA_SY',
'CICPCCADACurbRamp5in_ADA_SF',
'CICPCCADACurbRamp8in_ADA_SF',
'CICLinearDetWarnTiles_ADA_SF',
'CICBitumMatPrimeCoat_ADA_GAL',
'CICCurbGutLess300_ADA_FT',
'CICCurbGutGreater300_ADA_FT',
'CICProtConcSealer_CoreLFT',
'CICVertCurbT4_ADA_FT',
'CICHMAN30Type1_ADA_SY',
'CICHMAN30Type2_ADA_SY',
'CICHMAN70Type1_ADA_SY',
'CICHMAN70Type2_ADA_SY',
'CICDrillGroutTieBars_ADA_EA',
'CICPaveRemove_ADASY',
'CICHMASurfRemVD_CoreSY',
'CIC7ftPCCBaseCourse_ADASY',
'CIC9ftPCCBaseCourse_ADASY',
'CICHMAN30HM_CoreTON',
'CICHMAN70HM_CoreTON',
'CICHMAN30HM_ADATON',
'CICHMAN70HM_ADATON']:
            fieldnamesred.append(n)
    print fieldnamesred
    arcpy.AddMessage("Located Project Data...")

    def iterator_for_feature():
        cursor = arcpy.SearchCursor(dataset_name, whereClause)
        row = cursor.next()
        while row:
            yield [getattr(row, col) for col in fieldnamesred]
            row = cursor.next()
        del row, cursor
    return fieldnamesred, iterator_for_feature()
    arcpy.AddMessage("List Populated...")


#Dump the calc data to an excel sheet
wb = openpyxl.Workbook(optimized_write=True)
ws = wb.create_sheet()

# Header and row
header, rows = header_and_iterator(dataset_name)
ws.append(header)
for row in rows:
    ws.append(row)

# Save Output
wb.save(scratchData + '\CICCalcOutput.xlsx')
arcpy.AddMessage("Saving Data...")

#Open the just created excel workbook with the extracted table and copy the sheet into the template workbook
import win32com.client as win32
excel = win32.gencache.EnsureDispatch('Excel.Application')

wb2 = excel.Workbooks.Open(scratchData + '\CICCalcOutput.xlsx')

excel.Visible = False

#Select Data and Copy
excel.Range("1:2").Select()
excel.Selection.Copy()
arcpy.AddMessage("Copying Data to Excel Template File...")

#Paste into the template in a new worksheet
wb1 = excel.Workbooks.Open(templatePath + '\CICTemplate.xlsx')

wb1.Sheets('DataSheet').Select()

activesheet = wb1.ActiveSheet

excel.Range("1:2").Select()
excel.Selection.PasteSpecial()

#Save as a new workbook
wb1.SaveAs(outputWB + "\CICCalc_" + cce + '_' + blockNo + '.xlsx')
arcpy.AddMessage("Saving Data New File..." + outputWB + "\CICCalc_" + cce + '_' + blockNo + '.xlsx')
excel.Application.Quit()


